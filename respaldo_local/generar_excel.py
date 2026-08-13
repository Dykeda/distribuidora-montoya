"""Genera un Excel legible a partir de una copia de la base de datos (.db).

No reemplaza el respaldo real (el .db es la fuente de verdad y lo que se
usaría para restaurar el sistema) — esto es una foto adicional, en texto
plano, para poder abrir y revisar la información sin nada técnico si el
servidor llegara a fallar.

Uso: generar_excel.py <ruta_al_db> <ruta_al_xlsx_salida>
"""
import sys

from openpyxl import Workbook
from openpyxl.styles import Font


def hoja(wb, nombre, encabezados, filas):
    ws = wb.create_sheet(nombre)
    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = Font(bold=True)
    for fila in filas:
        ws.append(fila)
    for i, encabezado in enumerate(encabezados, start=1):
        ancho = max(len(str(encabezado)), 12)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = ancho + 4


def generar(ruta_db, ruta_salida):
    sys.path.insert(0, r"C:\Users\Kevin P\Documents\distribuidora-montoya")
    from app import create_app
    from models import (
        Producto,
        Compra,
        SalidaCamion,
        RetornoCamionDetalle,
        FacturaCartera,
        Gasto,
        CanjeDescuentoDetalle,
        AjusteCredito,
        VentaBodegaDetalle,
    )
    from services.ventas import cargado_por_producto

    app = create_app(config_overrides={"SQLALCHEMY_DATABASE_URI": "sqlite:///" + ruta_db})

    wb = Workbook()
    wb.remove(wb.active)

    with app.app_context():
        productos = Producto.query.order_by(Producto.nombre).all()
        hoja(
            wb, "Productos",
            ["Nombre", "Categoría", "Unidades/caja", "Precio actual (caja)", "Descuento ref. %", "Activo"],
            [
                [p.nombre, p.categoria or "", p.unidades_por_caja, p.precio_caja_actual() or 0,
                 p.tasa_descuento_referencia, "Sí" if p.activo else "No"]
                for p in productos
            ],
        )

        compras = Compra.query.order_by(Compra.fecha.desc()).all()
        filas_compras = []
        for c in compras:
            for d in c.detalles:
                filas_compras.append([
                    c.fecha, c.numero_factura or "", d.producto.nombre,
                    d.cantidad_comprada_unidades, d.costo_linea, d.tasa_descuento_aplicada,
                    d.credito_generado,
                ])
        hoja(
            wb, "Compras",
            ["Fecha", "Factura", "Producto", "Cantidad (unid.)", "Costo", "Tasa descuento %", "Crédito generado"],
            filas_compras,
        )

        salidas = SalidaCamion.query.order_by(SalidaCamion.fecha.desc()).all()
        filas_salidas, filas_retornos = [], []
        for s in salidas:
            for pid, cant in cargado_por_producto(s).items():
                producto = next((p for p in productos if p.id == pid), None)
                filas_salidas.append([s.fecha, producto.nombre if producto else pid, cant])
            if s.retorno:
                for d in s.retorno.detalles:
                    filas_retornos.append([s.fecha, s.retorno.fecha, d.producto.nombre, d.cantidad_unidades])
        hoja(wb, "Camion Salidas", ["Fecha salida", "Producto", "Cantidad cargada (unid.)"], filas_salidas)
        hoja(
            wb, "Camion Retornos",
            ["Fecha salida", "Fecha retorno", "Producto", "Cantidad regresada (unid.)"],
            filas_retornos,
        )

        facturas = FacturaCartera.query.order_by(FacturaCartera.fecha.desc()).all()
        hoja(
            wb, "Cartera",
            ["Cliente", "Fecha factura", "Ruta (fecha salida)", "Monto", "Estado", "Fecha pago", "Notas"],
            [
                [f.cliente, f.fecha, f.salida.fecha if f.salida else "Deuda anterior",
                 f.monto, f.estado, f.fecha_pago or "", f.notas or ""]
                for f in facturas
            ],
        )

        gastos = Gasto.query.order_by(Gasto.fecha.desc()).all()
        hoja(
            wb, "Gastos",
            ["Fecha", "Categoría", "Tipo", "Monto", "Notas"],
            [[g.fecha, g.categoria.nombre, g.categoria.tipo, g.monto, g.notas or ""] for g in gastos],
        )

        canjes = CanjeDescuentoDetalle.query.all()
        hoja(
            wb, "Descuentos Canjes",
            ["Fecha", "Producto", "Cantidad (unid.)", "Valor usado"],
            [[c.canje.fecha, c.producto.nombre, c.cantidad_unidades, c.valor_usado] for c in canjes],
        )

        ajustes = AjusteCredito.query.order_by(AjusteCredito.fecha.desc()).all()
        hoja(
            wb, "Descuentos Ajustes",
            ["Fecha", "Monto", "Notas"],
            [[a.fecha, a.monto, a.notas or ""] for a in ajustes],
        )

        ventas_bodega = VentaBodegaDetalle.query.all()
        hoja(
            wb, "Venta Bodega",
            ["Fecha", "Producto", "Cantidad (unid.)", "Valor"],
            [[v.venta.fecha, v.producto.nombre, v.cantidad_unidades, v.valor] for v in ventas_bodega],
        )

    wb.save(ruta_salida)
    print(f"Excel generado en: {ruta_salida}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Uso: generar_excel.py <ruta_al_db> <ruta_al_xlsx_salida>")
        raise SystemExit(1)
    generar(sys.argv[1], sys.argv[2])
