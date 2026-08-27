"""Detecta compras donde Postobón aplicó menos descuento del acordado por producto
(tasa_descuento_referencia), para poder reclamárselo."""
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import func, or_

from extensions import db
from models import CompraDetalle, Compra, Producto, AjustePostobon, Proveedor

# Fecha de arranque para sumas "históricas" (todo lo que haya, no acotado a un mes).
DESDE_SIEMPRE = date(2000, 1, 1)


def listar_faltantes(fecha_inicio, fecha_fin):
    """Líneas de compra en el período donde el % de descuento aplicado fue menor al
    acordado por producto -- para reclamarle a Postobón. Solo considera compras de
    Postobón (otros proveedores tienen sus propios acuerdos de descuento, distintos a la
    tasa_descuento_referencia guardada por producto) con número de factura (una factura
    real que se le puede reclamar) -- un ingreso de inventario físico sin factura (ej. un
    conteo de bodega) no aplica. Una compra sin proveedor asignado (dato viejo, de antes
    de que existiera este campo) se trata como Postobón, el caso histórico normal."""
    detalles = (
        CompraDetalle.query.join(Compra)
        .join(Producto)
        .outerjoin(Proveedor, Compra.proveedor_id == Proveedor.id)
        .filter(Compra.fecha >= fecha_inicio, Compra.fecha <= fecha_fin)
        .filter(Compra.numero_factura.isnot(None))
        .filter(CompraDetalle.es_descuento.is_(False))
        .filter(or_(Proveedor.id.is_(None), Proveedor.es_postobon.is_(True)))
        .order_by(Compra.fecha.desc())
        .all()
    )
    filas = []
    for d in detalles:
        diferencia_pct = round(d.producto.tasa_descuento_referencia - d.tasa_descuento_aplicada, 1)
        if diferencia_pct <= 0:
            continue  # sin faltante, o aplicó igual o más de lo esperado
        monto_faltante = round(d.costo_linea * diferencia_pct / 100)
        filas.append({
            "compra": d.compra,
            "detalle": d,
            "producto": d.producto,
            "tasa_esperada": d.producto.tasa_descuento_referencia,
            "tasa_aplicada": d.tasa_descuento_aplicada,
            "diferencia_pct": diferencia_pct,
            "monto_faltante": monto_faltante,
        })
    return filas


def listar_faltantes_de_compra(compra_id):
    """Igual que listar_faltantes(), pero para una sola compra (factura) puntual. Si esa
    compra es de otro proveedor (no Postobón, y no un dato viejo sin proveedor asignado),
    devuelve una lista vacía -- no hay nada que reclamarle a Postobón por una factura de
    otro proveedor."""
    detalles = (
        CompraDetalle.query.join(Compra)
        .join(Producto)
        .outerjoin(Proveedor, Compra.proveedor_id == Proveedor.id)
        .filter(Compra.id == compra_id)
        .filter(Compra.numero_factura.isnot(None))
        .filter(CompraDetalle.es_descuento.is_(False))
        .filter(or_(Proveedor.id.is_(None), Proveedor.es_postobon.is_(True)))
        .all()
    )
    filas = []
    for d in detalles:
        diferencia_pct = round(d.producto.tasa_descuento_referencia - d.tasa_descuento_aplicada, 1)
        if diferencia_pct <= 0:
            continue
        monto_faltante = round(d.costo_linea * diferencia_pct / 100)
        filas.append({
            "compra": d.compra,
            "detalle": d,
            "producto": d.producto,
            "tasa_esperada": d.producto.tasa_descuento_referencia,
            "tasa_aplicada": d.tasa_descuento_aplicada,
            "diferencia_pct": diferencia_pct,
            "monto_faltante": monto_faltante,
        })
    return filas


def listar_faltantes_agrupados(fecha_inicio, fecha_fin):
    """Igual que listar_faltantes(), pero agrupado por factura -- cada grupo trae su
    propia lista de líneas y su subtotal, para que el informe no se vea suelto."""
    grupos_por_compra = {}
    for f in listar_faltantes(fecha_inicio, fecha_fin):
        compra = f["compra"]
        grupo = grupos_por_compra.setdefault(compra.id, {"compra": compra, "lineas": [], "subtotal": 0})
        grupo["lineas"].append(f)
        grupo["subtotal"] += f["monto_faltante"]

    grupos = list(grupos_por_compra.values())
    grupos.sort(key=lambda g: g["compra"].fecha, reverse=True)
    return grupos


def construir_workbook_faltantes(fecha_inicio, fecha_fin):
    wb = Workbook()
    ws = wb.active
    ws.title = "Faltantes de descuento"

    encabezados = ["Factura", "Fecha", "Producto", "Notas", "Cantidad", "Costo", "% esperado", "% aplicado", "Diferencia %", "Monto faltante"]
    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = Font(bold=True)

    total_general = 0
    for grupo in listar_faltantes_agrupados(fecha_inicio, fecha_fin):
        compra = grupo["compra"]
        for f in grupo["lineas"]:
            ws.append([
                compra.numero_factura, compra.fecha, f["producto"].nombre, f["detalle"].notas or "",
                f["detalle"].cantidad_comprada_unidades, f["detalle"].costo_linea,
                f["tasa_esperada"], f["tasa_aplicada"], f["diferencia_pct"], f["monto_faltante"],
            ])
        ws.append(["", "", "", "", "", "", "", "", "Subtotal factura", grupo["subtotal"]])
        for celda in ws[ws.max_row]:
            celda.font = Font(bold=True)
        total_general += grupo["subtotal"]

    ws.append(["", "", "", "", "", "", "", "", "TOTAL", total_general])
    for celda in ws[ws.max_row]:
        celda.font = Font(bold=True)

    for i, encabezado in enumerate(encabezados, start=1):
        ancho = max(len(str(encabezado)), 12)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = ancho + 4

    return wb


def listar_ajustes():
    return AjustePostobon.query.order_by(AjustePostobon.fecha.desc(), AjustePostobon.id.desc()).all()


def total_ajustes_hasta(fecha_corte):
    total = (
        db.session.query(func.coalesce(func.sum(AjustePostobon.monto), 0))
        .filter(AjustePostobon.fecha <= fecha_corte)
        .scalar()
    )
    return round(total)


def total_pendiente_acumulado(fecha_corte=None):
    """Saldo pendiente de Postobón de todo el tiempo hasta la fecha de corte: la suma de
    todos los faltantes detectados en las facturas + los ajustes manuales (deuda anterior
    a esta pantalla, o abonos que Postobón haya hecho). No se acota por mes -- es el
    saldo que se le puede reclamar hoy."""
    fecha_corte = fecha_corte or date.today()
    faltante_historico = sum(f["monto_faltante"] for f in listar_faltantes(DESDE_SIEMPRE, fecha_corte))
    return faltante_historico + total_ajustes_hasta(fecha_corte)


def construir_workbook_faltantes_de_compra(compra_id):
    wb = Workbook()
    ws = wb.active
    ws.title = "Faltantes de descuento"

    encabezados = ["Factura", "Fecha", "Producto", "Notas", "Cantidad", "Costo", "% esperado", "% aplicado", "Diferencia %", "Monto faltante"]
    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = Font(bold=True)

    filas = listar_faltantes_de_compra(compra_id)
    total = 0
    for f in filas:
        compra = f["compra"]
        ws.append([
            compra.numero_factura, compra.fecha, f["producto"].nombre, f["detalle"].notas or "",
            f["detalle"].cantidad_comprada_unidades, f["detalle"].costo_linea,
            f["tasa_esperada"], f["tasa_aplicada"], f["diferencia_pct"], f["monto_faltante"],
        ])
        total += f["monto_faltante"]

    ws.append(["", "", "", "", "", "", "", "", "TOTAL", total])
    for celda in ws[ws.max_row]:
        celda.font = Font(bold=True)

    for i, encabezado in enumerate(encabezados, start=1):
        ancho = max(len(str(encabezado)), 12)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = ancho + 4

    return wb
