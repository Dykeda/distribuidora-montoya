"""Genera un Excel legible con toda la información del negocio, una hoja por tipo de
dato. Usado tanto por la descarga directa desde la web (routes/reportes.py) como por el
script de respaldo local (respaldo_local/generar_excel.py) — ahí es la única fuente de
esta lógica, para no mantenerla dos veces.

Además de las tablas en bruto, incluye una hoja "Resumen" y "Descuento por producto"
con los mismos cálculos que ya existen en la app (descuento contabilizado, cartera
pendiente, saldo de caja, venta implícita en dinero) — así el Excel sirve de verdad como
respaldo legible, no solo como un volcado de tablas."""
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font

from models import (
    Producto,
    Compra,
    CompraDetalle,
    SalidaCamion,
    FacturaCartera,
    Gasto,
    VentaBodegaDetalle,
)
from services.ventas import cargado_por_producto, venta_por_salida, ventas_en_periodo
from services.descuentos import (
    total_descuento_periodo,
    total_descuento_total_hasta,
    rendimiento_por_producto,
)
from services.cartera import total_pendiente
from services.caja import saldo_acumulado as saldo_caja_acumulado
from services.reportes import compra_total_periodo
from services.gastos import total_gastos_periodo

DESDE_SIEMPRE = date(2000, 1, 1)


def _hoja(wb, nombre, encabezados, filas):
    ws = wb.create_sheet(nombre)
    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = Font(bold=True)
    for fila in filas:
        ws.append(fila)
    for i, encabezado in enumerate(encabezados, start=1):
        ancho = max(len(str(encabezado)), 12)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = ancho + 4


def construir_workbook():
    """Arma el Excel a partir de los datos actuales. Debe llamarse dentro de un
    contexto de aplicación Flask con acceso a la base de datos."""
    productos = Producto.query.order_by(Producto.nombre).all()
    hoy = date.today()

    wb = Workbook()
    wb.remove(wb.active)

    compra_historica = compra_total_periodo(DESDE_SIEMPRE, hoy)
    venta_historica = ventas_en_periodo(DESDE_SIEMPRE, hoy)
    descuento_contabilizado_historico = total_descuento_total_hasta(hoy)
    cartera = total_pendiente(hoy)
    caja = saldo_caja_acumulado(hoy)
    pct_descuento = (
        round(descuento_contabilizado_historico / compra_historica["dinero"] * 100, 1)
        if compra_historica["dinero"] > 0 else 0.0
    )
    gastos_negocio_historico = total_gastos_periodo(None, hoy, tipo="negocio")
    ganancia_neta_historica = descuento_contabilizado_historico - gastos_negocio_historico

    _hoja(
        wb, "Resumen",
        ["Concepto", "Valor (a hoy)"],
        [
            ["Fecha de este resumen", hoy],
            ["Compra total (histórico)", compra_historica["dinero_con_iva"]],
            ["Venta total (histórico)", venta_historica["total"]],
            ["Descuento contabilizado (histórico)", descuento_contabilizado_historico],
            ["Cartera pendiente por cobrar", cartera],
            ["Saldo de caja acumulado", caja],
            ["% de descuento promedio (histórico)", pct_descuento],
            ["Ganancia neta del negocio (histórico)", ganancia_neta_historica],
        ],
    )

    _hoja(
        wb, "Descuento por producto",
        ["Producto", "Cajas (histórico)", "Descuento contabilizado (histórico)"],
        [
            [r["producto"].nombre, r["cajas"] if r["cajas"] is not None else "", r["descuento_contabilizado"]]
            for r in rendimiento_por_producto(DESDE_SIEMPRE, hoy)
        ],
    )

    _hoja(
        wb, "Productos",
        ["Nombre", "Categoría", "Unidades/caja", "Precio actual (caja)", "Descuento ref. %", "Activo"],
        [
            [p.nombre, p.categoria or "", p.unidades_por_caja, p.precio_caja_actual() or 0,
             p.tasa_descuento_referencia, "Sí" if p.activo else "No"]
            for p in productos
        ],
    )

    compras = Compra.query.order_by(Compra.fecha.desc()).all()
    filas_compras = []
    for c in compras:
        for d in c.detalles:
            filas_compras.append([
                c.fecha, c.numero_factura or "", d.producto.nombre, d.notas or "",
                d.cantidad_comprada_unidades, d.costo_linea, d.tasa_descuento_aplicada,
                "Sí" if d.es_descuento else "No", d.porcentaje_iva, d.valor_iva,
                c.proveedor.nombre if c.proveedor else "Postobón",
            ])
    _hoja(
        wb, "Compras",
        ["Fecha", "Factura", "Producto", "Notas", "Cantidad (unid.)", "Costo", "Tasa descuento %", "Es descuento", "% IVA", "Valor IVA", "Proveedor"],
        filas_compras,
    )

    salidas = SalidaCamion.query.order_by(SalidaCamion.fecha.desc()).all()
    filas_salidas, filas_ventas = [], []
    for s in salidas:
        for pid, cant in cargado_por_producto(s).items():
            producto = next((p for p in productos if p.id == pid), None)
            filas_salidas.append([s.fecha, producto.nombre if producto else pid, cant])
        if s.retorno:
            for v in venta_por_salida(s.id) or []:
                filas_ventas.append([
                    s.fecha, s.retorno.fecha, v["producto"].nombre,
                    v["cantidad_vendida"], v["precio_usado"], v["valor"],
                ])
    _hoja(wb, "Camion Salidas", ["Fecha salida", "Producto", "Cantidad cargada (unid.)"], filas_salidas)
    _hoja(
        wb, "Camion Ventas",
        ["Fecha salida", "Fecha retorno", "Producto", "Cantidad vendida (unid.)", "Precio usado", "Valor"],
        filas_ventas,
    )

    facturas = FacturaCartera.query.order_by(FacturaCartera.fecha.desc()).all()
    _hoja(
        wb, "Cartera",
        ["Cliente", "Fecha factura", "Ruta (fecha salida)", "Monto", "Abonado", "Saldo", "Estado", "Fecha pago", "Notas"],
        [
            [f.cliente.nombre, f.fecha, f.salida.fecha if f.salida else "Deuda anterior",
             f.monto, f.total_abonado, f.saldo_pendiente, f.estado, f.fecha_pago or "", f.notas or ""]
            for f in facturas
        ],
    )

    gastos = Gasto.query.order_by(Gasto.fecha.desc()).all()
    _hoja(
        wb, "Gastos",
        ["Fecha", "Categoría", "Tipo", "Monto", "Notas"],
        [[g.fecha, g.categoria.nombre, g.categoria.tipo, g.monto, g.notas or ""] for g in gastos],
    )

    descuentos = (
        CompraDetalle.query.join(Compra)
        .filter(CompraDetalle.es_descuento.is_(True))
        .order_by(Compra.fecha.desc())
        .all()
    )
    _hoja(
        wb, "Descuentos",
        ["Fecha", "Factura", "Producto", "Cantidad (unid.)", "Costo", "Proveedor"],
        [
            [d.compra.fecha, d.compra.numero_factura or "", d.producto.nombre,
             d.cantidad_comprada_unidades, d.costo_linea,
             d.compra.proveedor.nombre if d.compra.proveedor else "Postobón"]
            for d in descuentos
        ],
    )

    ventas_bodega = VentaBodegaDetalle.query.all()
    _hoja(
        wb, "Venta Bodega",
        ["Fecha", "Producto", "Cantidad (unid.)", "Valor"],
        [[v.venta.fecha, v.producto.nombre, v.cantidad_unidades, v.valor] for v in ventas_bodega],
    )

    return wb
