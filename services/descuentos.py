"""Contabilización del "descuento en producto" que trae cada factura de Postobón:
líneas de compra que representan esa parte (marcadas con CompraDetalle.es_descuento),
agrupadas por factura. No es un crédito acumulable ni algo canjeable — es solo un
registro histórico, total y por compra, de cuánto ha sido ese descuento."""
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import func

from extensions import db
from models import CompraDetalle, Compra, Producto


def total_descuento_periodo(fecha_inicio, fecha_fin):
    total = (
        db.session.query(func.coalesce(func.sum(CompraDetalle.costo_linea), 0))
        .join(Compra, CompraDetalle.compra_id == Compra.id)
        .filter(Compra.fecha >= fecha_inicio, Compra.fecha <= fecha_fin)
        .filter(CompraDetalle.es_descuento.is_(True))
        .scalar()
    )
    return round(total)


def total_descuento_total_hasta(fecha_corte):
    total = (
        db.session.query(func.coalesce(func.sum(CompraDetalle.costo_linea), 0))
        .join(Compra, CompraDetalle.compra_id == Compra.id)
        .filter(Compra.fecha <= fecha_corte)
        .filter(CompraDetalle.es_descuento.is_(True))
        .scalar()
    )
    return round(total)


def listar_descuentos_agrupados(fecha_inicio, fecha_fin):
    """Líneas marcadas como descuento en el período, agrupadas por factura, cada grupo
    con su propio subtotal."""
    detalles = (
        CompraDetalle.query.join(Compra)
        .join(Producto)
        .filter(Compra.fecha >= fecha_inicio, Compra.fecha <= fecha_fin)
        .filter(CompraDetalle.es_descuento.is_(True))
        .order_by(Compra.fecha.desc())
        .all()
    )

    grupos_por_compra = {}
    for d in detalles:
        compra = d.compra
        grupo = grupos_por_compra.setdefault(compra.id, {"compra": compra, "lineas": [], "subtotal": 0})
        grupo["lineas"].append({"detalle": d, "producto": d.producto})
        grupo["subtotal"] += d.costo_linea

    grupos = list(grupos_por_compra.values())
    grupos.sort(key=lambda g: g["compra"].fecha, reverse=True)
    return grupos


def rendimiento_por_producto(fecha_inicio, fecha_fin):
    """Cuánto descuento contabilizado (líneas marcadas) generó cada producto en el
    período, de mayor a menor."""
    filas = (
        db.session.query(
            Producto.nombre,
            func.coalesce(func.sum(CompraDetalle.costo_linea), 0).label("descuento_contabilizado"),
        )
        .join(Compra, CompraDetalle.compra_id == Compra.id)
        .join(Producto, CompraDetalle.producto_id == Producto.id)
        .filter(Compra.fecha >= fecha_inicio, Compra.fecha <= fecha_fin)
        .filter(CompraDetalle.es_descuento.is_(True))
        .group_by(Producto.id)
        .all()
    )

    resultado = [
        {"producto": nombre, "descuento_contabilizado": round(valor)}
        for nombre, valor in filas
    ]
    resultado.sort(key=lambda f: f["descuento_contabilizado"], reverse=True)
    return resultado


def construir_workbook_descuentos(fecha_inicio, fecha_fin):
    wb = Workbook()
    ws = wb.active
    ws.title = "Descuentos"

    encabezados = ["Factura", "Fecha", "Producto", "Cantidad", "Costo"]
    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = Font(bold=True)

    total_general = 0
    for grupo in listar_descuentos_agrupados(fecha_inicio, fecha_fin):
        compra = grupo["compra"]
        for f in grupo["lineas"]:
            ws.append([
                compra.numero_factura or "-", compra.fecha, f["producto"].nombre,
                f["detalle"].cantidad_comprada_unidades, f["detalle"].costo_linea,
            ])
        ws.append(["", "", "", "Subtotal factura", grupo["subtotal"]])
        for celda in ws[ws.max_row]:
            celda.font = Font(bold=True)
        total_general += grupo["subtotal"]

    ws.append(["", "", "", "TOTAL", total_general])
    for celda in ws[ws.max_row]:
        celda.font = Font(bold=True)

    for i, encabezado in enumerate(encabezados, start=1):
        ancho = max(len(str(encabezado)), 12)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = ancho + 4

    return wb
